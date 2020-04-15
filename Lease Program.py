#MiCRO
import openpyxl
    
book = openpyxl.load_workbook('Lease Database.xlsx')
sheet = book.active

able = 2 
count = input ("How many customers? > ")
count = (int(count) + 2)
while str(able) != str(count):
	alpha =  sheet['F' + str(able)] #FIRST NAME
	charlie =  sheet['H' + str(able)] #PHONE 
	echo =  sheet['B' + str(able)] #LEASE ID
	foxtrot =  sheet['A' + str(able)] #VIN
	golf =  sheet['C' + str(able)] #YEAR
	hotel =  sheet['D' + str(able)] #MAKE
	india =  sheet['E' + str(able)] #MODEL
	kilo =  sheet['J' + str(able)] #LEASE START
	lima =  sheet['K' + str(able)] #LEASE END
	mike =  sheet['N' + str(able)] #TERM
	november =  sheet['O' + str(able)] #PAYMENT
	oscar =  sheet['P' + str(able)] #FREQUENCY
	papa =  sheet['Q' + str(able)] #MONTHLY EQUIVALENT
	quebec =  sheet['W' + str(able)] #RESIDUAL
	
	romeo = str(alpha.value).split()[0]
	
	wb = openpyxl.load_workbook('Lease Template.xlsx')
	ws = wb.active 
	ws2 = wb.copy_worksheet(ws)
	ws2.title = (romeo)	
	wb.save("Lease Template.xlsx")
		
	workbook = openpyxl.load_workbook('Lease Template.xlsx')
	worksheet = workbook[romeo]	
	
	worksheet.cell(row=5, column=2).value = (alpha.value)
	worksheet.cell(row=6, column=2).value = (charlie.value)
	worksheet.cell(row=8, column=2).value = (echo.value)
	worksheet.cell(row=9, column=2).value = (foxtrot.value)
	worksheet.cell(row=10, column=2).value = (golf.value)
	worksheet.cell(row=11, column=2).value = (hotel.value)
	worksheet.cell(row=12, column=2).value = (india.value)
	worksheet.cell(row=6, column=6).value = (kilo.value)[:8]
	worksheet.cell(row=7, column=6).value = (lima.value)[:8]
	worksheet.cell(row=8, column=6).value = (mike.value)
	worksheet.cell(row=9, column=6).value = (november.value)
	worksheet.cell(row=10, column=6).value = (oscar.value)
	worksheet.cell(row=11, column=6).value = (papa.value)
	worksheet.cell(row=12, column=6).value = (quebec.value)
	
	workbook.save('Lease Template.xlsx')
	
	print (str(able) + "'/" + str(count))
	
	able += 1
	 

